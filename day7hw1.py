import openpyxl as xl
exf = xl.load_workbook('c:\\dd\\itx.xlsx')
aws = exf.active
tot = 0
for i in aws.rows:
    index = i[0].row
    name = i[0].value
    salary = i[1].value
    
    tot += salary 
    avg = tot / 5

    aws.cell(row = 6, column = 1).value = 'avg'
    aws.cell(row = 6, column = 2).value = avg

    print('{}{}'.format(name, salary))

exf.save('c:\\dd\\outitx.xlsx')
exf.close()
